from selenium import webdriver
from time import sleep
import requests
import openpyxl
count_find = 0
start_index = 2

file = input("Enter path: ") # воодим путь к входному файлу

def get_data(tin):
	global count_find
	options = webdriver.ChromeOptions() # задание стандартных настроек для chromdriver

	options.add_argument('headless') # добовляем аргумент для сокрытия окна chromdriver 
	options.add_argument('--disable-gpu') # добовляем аргумент для отключения вовлеченности графики в процесс

	path = "chromedriver.exe" #название файла

	browser = webdriver.Chrome(executable_path=path, options=options) # запуск chromedriver

	url = "https://my2.soliq.uz/main/info/personal?searchtin=" + str(tin) # ссылка на страницу
	api = "https://my2.soliq.uz/main/info/personal/data?tin=" + str(tin) # ссылка на апи

	data = [] # создаем пустой массив
	
	wb = openpyxl.load_workbook("res.xlsx")
	work_sheet = wb.active


	try:
		response = requests.get(api).json() # выполняем запрос к апи

	
		data.append(response["data"]["tin"]) # получаем ИНН
		

		browser.get(url) # получаем страницу

		browser.execute_script("window.scrollTo(0, 600)") # запускаем скрипт для прокручивания до данных

		sleep(5) # ожидаем 5 секунд для загрузки всех данных

		#time.sleep(10)

		tin_face = browser.find_element_by_css_selector("#nameinfo p").text

		# У физ лиц нет имени организции, если физ лицо, то ставим -
		if tin_face != "Физическое лицо":
			data.append(response["data"]["name"]) # пролучаем имя ЮЛ
		else:
			data.append("-")


		data.append(tin_face) # статус инн

		try:
			data.append(browser.find_element_by_css_selector("#ndsStatus").text) # является плательщиком НДС или нет
		except:
			data.append("-")
		try:
			data.append(browser.find_element_by_css_selector("#debtorStatus").text) # должник или нет
		except:
			data.append("-")

		if tin_face != "Физическое лицо": #Физ лицо не показывает статус банкрота на сайте, проверяем это
			data.append(browser.find_element_by_css_selector("#bankrotStatus").text) # является банкротом или нет


		#У физических лиц по АПИ есть только адрес, записываем только его
		if tin_face == "Физическое лицо":
			data.append("-")
			data.append("-")
			try:
				data.append(response["data"]["address"]) # адрес
			except:
				data.append("-")
		else:
			try:
				data.append(response["data"]["ns1Name"]) # вид деятельности
			except:
				data.append("-")
			try:
				data.append(response["data"]["address"]) # адрес
			except:
				data.append("-")
			try:
				data.append(response["data"]["regDate"]) # дата регистрации
			except:
				data.append("-")
			try:
				data.append(response["data"]["regNum"]) # номер регестрации
			except:
				data.append("-")
			try:
				data.append(response["data"]["nc2Name"]) # ОКПО
			except:
				data.append("-")
			try:
				data.append(response["data"]["nc5Name"]) # СОАТО
			except:
				data.append("-")
			try:
				data.append(response["data"]["nc4Name"]) # ОПФ
			except:
				data.append("-")
			try:
				data.append(response["data"]["nc6Name"]) # ОКЭД
			except:
				data.append("-")
			try:
				data.append(response["data"]["nc1Name"]) # СООГУ
			except:
				data.append("-")
			
			
		print(url) # выводим ссылку на страницу
		
		count_find += 1 # увеличиваем количество найденых ЮЛ

		# записываем в файл полученный единичный ИНН

		print(data)
		work_sheet.append(data)
		wb.save(filename="res.xlsx")
		browser.quit() # закрываем браузер

	except Exception as e: # добавление ИНН и пустых строк, как знак того, что данного ИНН не обнаружено на сайте
		#Чтобы ИНН был в первой строчке, сделаем первый элемент массива ИННом
		if(len(data) > 0):
			data[0] = str(tin)
		else:
			data.append(str(tin)) #записываем не найденный ИНН
		for _ in range(1,15): #заполняем остальные поля -
			data.append("-")
		work_sheet.append(data)
		wb.save(filename="res.xlsx")
		print(url + ' НЕ НАЙДЕН') # выводим ссылку и " не найден"
		browser.quit() # закрываем браузер


wb = openpyxl.Workbook() # создаем файл xlsx 
sheet = wb.active # актививруем его
sheet["A1"] = "ИНН" # добавляем заголовки
sheet["B1"] = "Наименование"
sheet["C1"] = "ЮЛ"
sheet["D1"] = "НДС"
sheet["E1"] = "Является ли должником"
sheet["F1"] = "Является ли банкротом"
sheet["G1"] = "Вид деятельности"
sheet["H1"] = "Адрес"
sheet["I1"] = "Дата регистрации"
sheet["J1"] = "Номер регистрации"
sheet["K1"] = "ОКПО"
sheet["L1"] = "СОАТО"
sheet["M1"] = "ОПФ"
sheet["N1"] = "ОКЭД"
sheet["O1"] = "СООГУ"
wb.save(filename="res.xlsx") # сохраняем файл 

try:
	wb = openpyxl.reader.excel.load_workbook(filename=file) # открываем файл для чтения
except:
	print("Входной файл не найден!!!")
	input("Нажмите enter для выхода")

wb.active = 0

sheet = wb.active

index = 2

tins = []
while sheet["A"+str(index)].value != None: # получаем все ИНН для поиска 
	tins.append(sheet["A"+str(index)].value)
	index += 1

for tin in tins: # обрабатываем каждый ИНН
	get_data(tin)


print('ИНН найдено:', str(count_find)) 
print('ЮЛ не найдено по ИНН:', str(len(tins) - count_find))
